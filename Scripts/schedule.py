import os
import math
import requests
import openpyxl
import openpyxl.styles.numbers
import pyexcel as xlsx_file
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
from configure import path

link = {
    'site': 'https://old.sevsu.ru/univers/shedule',
    'schedule': 'https://old.sevsu.ru/images/raspis/2021-2022/2021/2%20SEMESTR/IITUTS/IITUTS_1kurs_25_04.xls'
}

file = {
    'legacy': 'LegacySchedule.xls',
    'converted': 'ConvertedSchedule.xlsx',
}


def have_legacy():
    return os.path.isfile(path.get('schedules') + file.get('legacy'))


def check_for_updates():
    result = requests.get(link.get('site')).content
    soup = BeautifulSoup(result, 'html.parser')
    file_urls = soup.find(class_="su-column-content").find_all("a")

    for file_url in file_urls:
        f = file_url.get("href")
        if 'IITUTS_1kurs' in str(f):
            template_link = link.get('schedule')[:69]
            current_name_file = link.get('schedule')[69:]
            acquired_name_file = str(f)[49:]

            if acquired_name_file != current_name_file:
                link['schedule'] = template_link + acquired_name_file
                return True

    return False


def download():
    downloaded_file = requests.get(link.get('schedule'))
    with open(path.get('schedules') + file.get('legacy'), 'wb') as legacy_file:
        legacy_file.write(downloaded_file.content)
    create_xlsx()


def create_xlsx():
    xlsx_file.save_book_as(file_name=path.get('schedules') + file.get('legacy'), dest_file_name=path.get('schedules') + file.get('converted'))


def create_new_schedule(group):
    schedule = openpyxl.load_workbook(path.get('schedules') + file.get('converted'))

    target_list = str(current_week()) + ' нед'
    if target_list in schedule.sheetnames:
        sheet = schedule[target_list]
    else:
        return False

    new_schedule = Workbook()
    new_sheet = new_schedule.active
    new_sheet.title = str(current_week()) + 'Неделя'

    class_shift = 24 + (group - 1) * 4
    date_shift = 22

    new_sheet['A1'] = str(sheet[get_column_letter(class_shift) + str(10)].value)
    new_sheet['A2'] = str(current_week()) + ' Неделя'

    weekday = 0
    weekday_text = ''
    pair = 8

    for r in range(11, 59):
        pair += 1
        if pair == 9:
            pair = 1
            weekday += 1

            weekday_text = str(sheet['U' + str(11 + (weekday - 1) * 8)].value)
            weekday_text = weekday_text.capitalize()
            new_sheet['A' + str(4 + (weekday - 1) * 9)].value = weekday_text

        for c in range(date_shift, date_shift + 2):
            cell = sheet[get_column_letter(c) + str(r)]
            new_sheet.cell(r - 7 + weekday, c - date_shift + 1, cell.value)

        for c in range(class_shift, class_shift + 4):
            cell = sheet[get_column_letter(c) + str(r)]
            cell.value = check_general_class(sheet, r, c - class_shift + 1 + 2, weekday_text, cell.value)
            if cell.value:
                new_sheet.cell(r - 7 + weekday, c - class_shift + 1 + 2, cell.value)

    formatting_new_schedule(new_sheet)
    new_schedule.save(path.get('schedules') + get_file_name(group) + '.xlsx')
    return True


def formatting_new_schedule(new_sheet):
    new_sheet.merge_cells('A1:F1')
    new_sheet.merge_cells('A2:F2')
    new_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    new_sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')

    col_weekdays = 'A'
    col_times = 'B'
    col_class_names = ['C', 'D']
    friday_row = range(41, 49)

    for r in range(4, 59):
        for c in range(1, 8):
            cell = new_sheet[get_column_letter(c) + str(r)]
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if cell.value:
                if get_column_letter(c) in col_weekdays:
                    if not new_sheet[get_column_letter(c + 1) + str(r)].value:
                        new_sheet.merge_cells(
                            get_column_letter(c) + str(r) + ':' + get_column_letter(c + 5) + str(r))

                if get_column_letter(c) in col_times:
                    cell.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[20]

                elif get_column_letter(c) in col_class_names:
                    cell.value = to_short_name(str(cell.value))

                    if not is_sub_class(str(cell.value)) or r in friday_row:
                        if 'ПУЛ' in str(cell.value):
                            new_sheet.merge_cells(
                                get_column_letter(c) + str(r) + ':' + get_column_letter(c + 2) + str(r + 7))
                            new_sheet.merge_cells(
                                get_column_letter(c + 3) + str(r) + ':' + get_column_letter(c + 3) + str(r + 7))
                        else:
                            new_sheet.merge_cells(
                                get_column_letter(c) + str(r) + ':' + get_column_letter(c + 1) + str(r))


def get_file_name(group):
    if group < 4:
        return 'ИСб-21-' + str(group) + '-о'
    return 'ПИб-21-1-о'


def check_general_class(sheet, row, colum, weekday_text, default_value):
    is1_class_name = sheet[get_column_letter(24) + str(row)]

    if is1_class_name.value:
        if get_column_letter(colum) in 'C':
            if weekday_text == 'Пятница' or 'ПУЛ' in str(is1_class_name.value) or 'язык' in str(is1_class_name.value):
                return is1_class_name.value

        elif get_column_letter(colum) in 'E':
            if weekday_text == 'Пятница':
                return 'Л'
            elif 'язык' in str(is1_class_name.value):
                return 'ПЗ'

    return default_value


def to_short_name(name):
    if 'Высшая' in name:
        return 'ВМ'
    elif 'Алгоритмизация' in name:
        return 'АИП'
    elif 'Физика. Электромагнетизм' in name:
        return 'Электромагнетизм'
    elif 'Основы Ф' in name:
        return 'Основы физики'
    elif 'Основы м' in name:
        return 'Основы математики'
    elif 'Инструментальные' in name:
        return 'Графика'
    elif 'История' in name:
        return 'История'
    elif 'Иностранный' in name:
        return 'Иностранный язык'
    else:
        return name


def is_sub_class(class_name):
    sub_classes = ['Графика']

    if class_name in sub_classes:
        return True
    return False


def current_week():
    difference_date = datetime.now() - datetime(2021, 9, 1)
    return math.ceil(difference_date.days / 7)
